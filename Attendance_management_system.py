import tkinter as tk
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import datetime

class Application(tk.Frame):
    def __init__(self, root):
        super().__init__(root, width=442, height=300, borderwidth=1, relief='groove')
        self.root = root
        self.pack()
        self.pack_propagate(False)
        self.create_widgets()

    def create_widgets(self):
        title_label = tk.Label(self, text='勤怠管理システム', font=('Yu Gothic UI', '24', 'normal'))
        title_label.place(x=0, y=0)

        id_label = tk.Label(self, text='社員ID', font=('Yu Gothic UI', '15', 'normal'))
        id_label.place(x=0, y=130)

        pass_label = tk.Label(self, text='パスワード', font=('Yu Gothic UI', '15', 'normal'))
        pass_label.place(x=200, y=130)

        self.id_entry = tk.Entry(self)
        self.id_entry.place(x=65, y=140)

        self.pass_entry = tk.Entry(self, show='●')
        self.pass_entry.place(x=280, y=140)

        attendance_btn = tk.Button(self, text='出勤', width=10)
        attendance_btn['command'] = self.attendance
        attendance_btn.place(x=130, y=250)

        leabing_btn = tk.Button(self, text='退勤', width=10)
        leabing_btn['command'] = self.leabing
        leabing_btn.place(x=250, y=250)

    def attendance(self):
        input_id = self.id_entry.get()
        input_pass = str(self.pass_entry.get())
        self.remove_error_message()

        if len(input_id) == 0 or len(input_pass) == 0:
            error_message = tk.Label(self, text='社員IDとパスワードを入力してください。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
            error_message.place(x=0, y=175)
        else:
            wb = load_workbook('C:\\python\\Excel\\社員一覧.xlsx')
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if str(row[0]) == input_id and str(row[2]) == input_pass:
                    wb_attendance = load_workbook('C:\\python\\Excel\\勤怠管理.xlsx')
                    sheet_attendance = wb_attendance.active
                    last_row = sheet_attendance.max_row

                    name = row[1]

                    for row in sheet_attendance.iter_rows(values_only=True):
                        if str(row[0]) == input_id and row[4] == '出勤':
                            error_message = tk.Label(self, text='既に出勤済みです。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
                            error_message.place(x=0, y=175)
                            break
                    else:
                        now = datetime.datetime.now()
                        date_str = now.strftime('%Y-%m-%d')
                        time_str = now.strftime('%H:%M:%S')
                        cell = sheet_attendance.cell(row=last_row + 1, column=1, value=input_id)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=2, value=name)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=3, value=date_str)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=4, value=time_str)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=5, value='出勤')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        wb_attendance.save('C:\\python\\Excel\\勤怠管理.xlsx')
                        success_message = tk.Label(self, text='出勤しました。', font=('Yu Gothic UI', '8', 'normal'))
                        success_message.place(x=0, y=175)
                    wb_attendance.close()
                    break
            else:
                error_message = tk.Label(self, text='社員IDとパスワードが正しくありません。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
                error_message.place(x=0, y=175)
            wb.close()

    def leabing(self):
        input_id = self.id_entry.get()
        input_pass = str(self.pass_entry.get())
        self.remove_error_message()

        if len(input_id) == 0 or len(input_pass) == 0:
            error_message = tk.Label(self, text='社員IDとパスワードを入力してください。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
            error_message.place(x=0, y=175)
        else:
            wb = load_workbook('C:\\python\\Excel\\社員一覧.xlsx')
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if str(row[0]) == input_id and str(row[2]) == input_pass:
                    print('一致')
                    wb_attendance = load_workbook('C:\\python\\Excel\\勤怠管理.xlsx')
                    sheet_attendance = wb_attendance.active
                    last_row = sheet_attendance.max_row
                    name = row[1]

                    for row in sheet_attendance.iter_rows(values_only=True):
                        if str(row[0]) == input_id and row[4] == '退勤':
                            error_message = tk.Label(self, text='既に退勤済みです。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
                            error_message.place(x=0, y=175)
                            break
                    else:
                        now = datetime.datetime.now()
                        date_str = now.strftime('%Y-%m-%d')
                        time_str = now.strftime('%H:%M:%S')
                        cell = sheet_attendance.cell(row=last_row + 1, column=1, value=input_id)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=2, value=name)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=3, value=date_str)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=4, value=time_str)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell = sheet_attendance.cell(row=last_row + 1, column=5, value='退勤')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        wb_attendance.save('C:\\python\\Excel\\勤怠管理.xlsx')
                        success_message = tk.Label(self, text='退勤しました。', font=('Yu Gothic UI', '8', 'normal'))
                        success_message.place(x=0, y=175)
                    wb_attendance.close()
                    break
            else:
                error_message = tk.Label(self, text='社員IDとパスワードが正しくありません。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
                error_message.place(x=0, y=175)
            wb.close()

    def remove_error_message(self):
        for widget in self.winfo_children():
            if(isinstance(widget, tk.Label) and widget.cget('text') == '社員IDとパスワードを入力してください。' or isinstance(widget, tk.Label) and widget.cget('text') == '社員IDとパスワードが正しくありません。' or isinstance(widget, tk.Label) and widget.cget('text') == '既に出勤済みです。' or isinstance(widget, tk.Label) and widget.cget('text') == '既に退勤済みです。' or isinstance(widget, tk.Label) and widget.cget('text') == '出勤しました。' or isinstance(widget, tk.Label) and widget.cget('text') == '退勤しました。'):
                widget.destroy()


root = tk.Tk()
root.title('勤怠管理システム')
root.geometry('442x300')
app = Application(root=root)
root.mainloop()
